from openpyxl import Workbook, load_workbook

# Load the excel workbook and copy the Slack IDs to a new worksheet named to the current day
def init_attendance(excel_file, date_today, num_rows):
    Attendance = {}
    wb = load_workbook(excel_file)
    ids = wb['Classlist']

    for row in ids.iter_rows(min_row=1, min_col=1, max_row=num_rows, max_col=1):
        for cell in row:
            id = ids.cell(row=cell.row, column=cell.column).value
            Attendance[id] = None

    ws = wb.create_sheet(date_today)

    return wb, ws, Attendance

# Update the attendance list and put check mark to the ID that is present
def record_attendance(Attendance, id):
    if id in Attendance:
        Attendance[id] = "✓"

    return Attendance

# Called before the close of attendance checking and save to the excel file
def save_attendance(wb, ws, Attendance, excel_file):
    ws.column_dimensions['A'].width = 20
    r,c=1,1

    for k in Attendance:
        ws.cell(r, c, k)
        ws.cell(r, c+1, Attendance[k])
        r = r+1

    wb.save(excel_file)

if __name__ == "__main__":
    # Used for testing the functions above
    wb, ws, att = init_attendance('attendance.xlsx', 'test', 15)
    att = record_attendance(att, 'U057BCVHWCQ')
    att = record_attendance(att, 'U05R1AXHDCM')
    att = record_attendance(att, 'U05R48QL4JE')
    save_attendance(wb, ws, att, 'attendance.xlsx')