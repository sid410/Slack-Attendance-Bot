from openpyxl import Workbook, load_workbook

def grades_dictionary(excel_file):
    wb = load_workbook(excel_file)
    details = wb['Classlist']
    student_dict = {}
    grade = []
    id = ""

    for row in details.iter_rows(min_row=2, min_col=1, max_row=42, max_col=10):
        for cell in row:

            val = details.cell(row=cell.row, column=cell.column).value

            if cell.column == 1:
                grade = []
                id = val
            else:
                grade.append(val)

            student_dict[id] = grade

    return student_dict

if __name__ == "__main__":
    show_dict = grades_dictionary('grades-test.xlsx')
    print(show_dict)